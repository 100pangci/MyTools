import os
import glob
import shutil
import copy
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


# ================= 核心辅助函数：完美复制单元格样式 =================
def copy_cell_style(source_cell, target_cell):
    """复制单元格的所有核心样式（字体、对齐、边框、填充、数字格式）"""
    if source_cell.has_style:
        if source_cell.font:
            target_cell.font = copy.copy(source_cell.font)
        if source_cell.alignment:
            target_cell.alignment = copy.copy(source_cell.alignment)
        if source_cell.border:
            target_cell.border = copy.copy(source_cell.border)
        if source_cell.fill:
            target_cell.fill = copy.copy(source_cell.fill)
        if source_cell.number_format:
            target_cell.number_format = copy.copy(source_cell.number_format)
        if source_cell.protection:
            target_cell.protection = copy.copy(source_cell.protection)


def merge_excel_template(input_folder, output_file):
    files = glob.glob(os.path.join(input_folder, "*.xlsx"))
    files = [f for f in files if os.path.basename(f) != os.path.basename(output_file) and not os.path.basename(f).startswith("~$")]

    if not files:
        print("❌ 未找到需要合并的 Excel 文件！")
        return

    print(f"正在以 {os.path.basename(files[0])} 为模板创建总表...")
    template_file = files[0]
    shutil.copy(template_file, output_file)

    try:
        wb_out = load_workbook(output_file)
        ws_out = wb_out['sheet1']

        max_col = ws_out.max_column
        source_col = max_col + 1
        ws_out.cell(row=1, column=source_col, value='Source_Filename')

        # 标记模板文件自己的数据源
        for r in range(2, ws_out.max_row + 1):
            ws_out.cell(row=r, column=source_col, value=os.path.basename(template_file))

        current_out_row = ws_out.max_row + 1

        for file in files[1:]:
            filename = os.path.basename(file)
            wb_in = load_workbook(file, data_only=True)
            if 'sheet1' not in wb_in.sheetnames:
                continue
            ws_in = wb_in['sheet1']

            for r in range(2, ws_in.max_row + 1):
                # 如果整行为空则跳过
                if all(ws_in.cell(row=r, column=c).value is None for c in range(1, max_col + 1)):
                    continue

                # 关键修复：复制原表的行高
                if r in ws_in.row_dimensions:
                    ws_out.row_dimensions[current_out_row].height = ws_in.row_dimensions[r].height

                for c in range(1, max_col + 1):
                    source_cell = ws_in.cell(row=r, column=c)
                    target_cell = ws_out.cell(row=current_out_row, column=c, value=source_cell.value)
                    copy_cell_style(source_cell, target_cell)

                # 记录文件名
                ws_out.cell(row=current_out_row, column=source_col, value=filename)
                current_out_row += 1

            wb_in.close()
            print(f"已追加合并: {filename}")

        wb_out.save(output_file)
        print(f"\n✅ 成功合并所有文件！高度、边框和格式已完美保留，保存为: {output_file}")

    except Exception as e:
        print(f"❌ 合并过程中出错: {e}")


def split_excel_template(input_file, output_folder):
    if not os.path.exists(output_folder):
        os.makedirs(output_folder)

    try:
        wb_in = load_workbook(input_file)
        ws_in = wb_in['sheet1']

        max_col = ws_in.max_column
        source_col_idx = None
        for c in range(1, max_col + 1):
            if ws_in.cell(row=1, column=c).value == 'Source_Filename':
                source_col_idx = c
                break

        if not source_col_idx:
            print("❌ 未找到 'Source_Filename' 列，无法拆分。")
            return

        file_row_mapping = {}
        for r in range(2, ws_in.max_row + 1):
            filename = ws_in.cell(row=r, column=source_col_idx).value
            if filename:
                if filename not in file_row_mapping:
                    file_row_mapping[filename] = []
                file_row_mapping[filename].append(r) 

        # 还原你的逻辑：基于总表制作模板（保留表头、列宽、筛选按钮等）
        wb_template = load_workbook(input_file)
        ws_template = wb_template['sheet1']
        ws_template.delete_cols(source_col_idx)
        
        # 删除第2行到末尾
        if ws_template.max_row >= 2:
            ws_template.delete_rows(2, ws_template.max_row)

        # ====================================================================
        # 🌟 核心修复：暴力扫除 delete_rows 遗留的“幽灵行”和“幽灵格式”
        # ====================================================================
        # 1. 清除底层残留的几万个行高记录（这是导致文件几十MB的元凶）
        ghost_rows = [r for r in ws_template.row_dimensions.keys() if r >= 2]
        for r in ghost_rows:
            del ws_template.row_dimensions[r]

        # 2. 清除底层残留的不可见空单元格对象
        ghost_cells = [coord for coord in ws_template._cells.keys() if coord[0] >= 2]
        for coord in ghost_cells:
            del ws_template._cells[coord]
            
        # 3. 修复筛选按钮的范围（防止下拉列表去扫描几万行空行）
        if ws_template.auto_filter.ref:
            max_col_letter = get_column_letter(ws_template.max_column)
            ws_template.auto_filter.ref = f"A1:{max_col_letter}1"
        # ====================================================================

        temp_template_path = "temp_blank_template.xlsx"
        wb_template.save(temp_template_path)

        for filename, row_indices in file_row_mapping.items():
            out_path = os.path.join(output_folder, filename)
            shutil.copy(temp_template_path, out_path)

            wb_out = load_workbook(out_path)
            ws_out = wb_out['sheet1']

            current_out_row = 2
            for original_row_idx in row_indices:
                if original_row_idx in ws_in.row_dimensions:
                    ws_out.row_dimensions[current_out_row].height = ws_in.row_dimensions[original_row_idx].height

                for c in range(1, source_col_idx):
                    source_cell = ws_in.cell(row=original_row_idx, column=c)
                    target_cell = ws_out.cell(row=current_out_row, column=c, value=source_cell.value)
                    copy_cell_style(source_cell, target_cell)

                current_out_row += 1
                
            # 🌟 修复：数据写完后，把筛选按钮的控制范围扩大到当前实际的行数
            if ws_out.auto_filter.ref:
                max_col_letter = get_column_letter(ws_out.max_column)
                ws_out.auto_filter.ref = f"A1:{max_col_letter}{ws_out.max_row}"

            wb_out.save(out_path)
            print(f"已拆分保存并还原格式: {filename}")

        if os.path.exists(temp_template_path):
            os.remove(temp_template_path)

        print(f"\n✅ 成功拆分！无用空行已彻底抹除，筛选按钮正常保留，保存在: {output_folder}")

    except Exception as e:
        print(f"❌ 拆分失败，错误: {e}")


if __name__ == "__main__":
    WORK_DIR = "."  
    MERGED_FILE = "合并结果总表.xlsx"  
    SPLIT_DIR = "./还原拆分文件"      

    print("请选择操作：\n1. 纵向合并文件（完美保留格式版）\n2. 拆分文件\n3. 退出")
    choice = input("请输入序号 (1/2/3): ")

    if choice == '1':
        print("-" * 30)
        merge_excel_template(WORK_DIR, MERGED_FILE)
    elif choice == '2':
        print("-" * 30)
        split_excel_template(MERGED_FILE, SPLIT_DIR)
    else:
        print("已退出。")